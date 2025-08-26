import serial
import time
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import norm, zscore, skew, kurtosis, anderson, probplot, kstest, jarque_bera
import pandas as pd
import seaborn as sns
import platform
import sys
import os
import statistics

port = 'COM11'
baudrate = 9600
max_len = 100
sensor_range = 500
version = 5

output_file_name = 'sensor_data(' + str(sensor_range) + ')mm(V' + str(version) + ').xlsx'

ser = serial.Serial(port, baudrate, timeout=1)
time.sleep(1)

data_list = []
cnt = 0
elapsed = 0

try:
    start_time = time.time()
    while cnt < max_len:
        if ser.in_waiting > 0:
            line = ser.readline().decode('utf-8').strip()
            try:
                value = int(line)
                if value == -1:
                    print("-1!")
                    continue
                data_list.append(value)
                cnt += 1
                elapsed = time.time() - start_time
                print(value)
                if cnt % 10 == 0:
                    print(f"{cnt/100*max_len}% done! sampling rate : ", cnt / elapsed)
            except ValueError:
                print(f"Ignored data: {line}")
except KeyboardInterrupt:
    print("Stopped")

ser.close()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SensorData"

ws['A1'] = "Index"
ws['B1'] = "Value"

for i, val in enumerate(data_list, start=1):
    ws.cell(row=i+1, column=1, value=i)
    ws.cell(row=i+1, column=2, value=val)

wb.save(output_file_name)
print("Done!")

###########################################################

class Logger(object):
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w", encoding="utf-8")
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        self.terminal.flush()
        self.log.flush()

file_name = output_file_name
excel_file = 'sensor_data(all)(V'+ str(version) + ').xlsx'
df = pd.read_excel(file_name)

col_name = df['Index']
data = df['Value']

base_name = os.path.splitext(file_name)[0]
log_file_name = base_name + "분석결과.txt"

sys.stdout = Logger(log_file_name)

if platform.system() == 'Windows':
    plt.rc('font', family='Malgun Gothic')
elif platform.system() == 'Darwin':
    plt.rc('font', family='AppleGothic')
else:
    plt.rc('font', family='NanumGothic')
plt.rcParams['axes.unicode_minus'] = False

z_scores = zscore(data)
mask = np.abs(z_scores) < 4
filtered_data = data[mask]

removed_count = len(data) - len(filtered_data)
removed_ratio = removed_count / len(data) * 100

def analyze_normality(arr, label):
    print(f"\n--- {label} ---")
    print(f"데이터 개수: {len(arr)}")
    
    mu, std = norm.fit(arr)
    print(f"평균(μ): {mu:.4f}, 표준편차(σ): {std:.4f}")
    
    skewness = skew(arr)
    kurt = kurtosis(arr, fisher=False)  # 일반 첨도 (정규=3)
    print(f"왜도(Skewness): {skewness:.4f}")
    print(f"첨도(Kurtosis): {kurt:.4f}")

    # Jarque-Bera 테스트
    jb_stat, jb_p = jarque_bera(arr)
    print(f"Jarque-Bera 통계량: {jb_stat:.4f}, p-value: {jb_p:.4g}")
    if jb_p > 0.05:
        print("  -> 귀무가설 채택 (정규분포 따름)")
    else:
        print("  -> 귀무가설 기각 (정규분포 따르지 않음)")
    return mu, std

fig, axs = plt.subplots(1, 2, figsize=(12,6))

probplot(data, dist="norm", plot=axs[0])
axs[0].set_title("① 이상치 제거 전 Q-Q Plot")
axs[0].grid(True)

probplot(filtered_data, dist="norm", plot=axs[1])
axs[1].set_title("② 이상치 제거 후 Q-Q Plot")
axs[1].grid(True)

plt.tight_layout()
plt.show()

print(f"총 데이터 수: {len(data)}")
print(f"이상치 제거 수: {removed_count} ({removed_ratio:.2f}%)")

analyze_normality(data, "이상치 제거 전 데이터")
Mu, Std = analyze_normality(filtered_data, "이상치 제거 후 데이터")

sys.stdout.log.close()
sys.stdout = sys.stdout.terminal

new_data = pd.DataFrame([{
    "실제 거리(mm)": sensor_range,
    "평균(mm)": round(Mu, 4),
    "표준편차(mm)": round(Std, 4)
}])

if os.path.exists(excel_file):
    old_df = pd.read_excel(excel_file)
    df = pd.concat([old_df, new_data], ignore_index = True)
else:
    df = new_data

df.to_excel(excel_file, index=False)
print("done2!")
