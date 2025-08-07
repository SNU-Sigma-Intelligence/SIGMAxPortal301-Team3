import serial
import time
import openpyxl

port = 'COM9'
baudrate = 115200
max_len = 1000
sensor_range = 400
version = 2

output_file_name = 'sensor_data(' + str(sensor_range) + ')mm(V' + str(version) + ').xlsx'

ser = serial.Serial(port, baudrate, timeout=1)
time.sleep(2)

data_list = []
cnt = 0

try:
    start_time = time.time()
    while cnt < max_len:
        if ser.in_waiting > 0:
            line = ser.readline().decode('utf-8').strip()
            try:
                value = int(line)
                data_list.append(value)
                cnt += 1
                elapsed = time.time() - start_time
                if cnt % 100 == 0:
                    print(f"{cnt/100}% done! sampling rate : ", cnt / elapsed)
            except ValueError:
                print(f"Ignored data: {line}")
except KeyboardInterrupt:
    print("Stopped")

ser.close()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SensorData"

ws['A1'] = "Index"
ws['B1'] = "Value"

for i, val in enumerate(data_list, start=1):
    ws.cell(row=i+1, column=1, value=i)
    ws.cell(row=i+1, column=2, value=val)

wb.save(output_file_name)
print("Done!")